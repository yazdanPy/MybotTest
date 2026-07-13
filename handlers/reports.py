# -*- coding: utf-8 -*-
"""Personal balance, paginated history (with delete), full group report, and Excel export."""

import io
import re

import jalali_utils as ju
import keyboards as kb
from balances import personal_balance, group_report
from handlers.common import get_active_user_or_warn, get_admin_or_warn, get_db
from telegram import Update
from telegram.ext import (
    ContextTypes,
    CommandHandler,
    MessageHandler,
    filters,
    CallbackQueryHandler,
)

HISTORY_PAGE_SIZE = 5


# --------------------------------------------------------------- my balance
async def balance_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await get_active_user_or_warn(update, context)
    if user is None:
        return
    db = get_db(context)
    lines_owe = []
    lines_owed = []
    for line in personal_balance(db, user["id"]):
        card = (
            ju.format_card_number_html(line.other_card)
            if line.other_card
            else "شماره کارت ثبت نشده"
        )
        if line.i_owe_them:
            lines_owe.append(
                f"  • به {line.other_name}: {ju.format_money(line.amount)}  (کارت: {card})"
            )
        else:
            lines_owed.append(
                f"  • از {line.other_name}: {ju.format_money(line.amount)}"
            )

    if not lines_owe and not lines_owed:
        text = "🎉 حسابت کاملاً صافه! نه به کسی بدهکاری، نه کسی بهت بدهکاره."
    else:
        parts = [f"📊 <b>وضعیت حساب {user['first_name']}</b>", ""]
        if lines_owe:
            parts.append("🔴 بدهکاری:")
            parts.extend(lines_owe)
            parts.append("")
        if lines_owed:
            parts.append("🟢 طلبکاری:")
            parts.extend(lines_owed)
        text = "\n".join(parts)

    await update.effective_message.reply_text(text, parse_mode="HTML")


# ------------------------------------------------------------------ history
def _format_expense_detail(db, expense: dict) -> str:
    lines = [
        f"<b>#{expense['id']} — {expense['description']}</b>",
    ]
    if expense.get("program_id"):
        prog = db.get_program(expense["program_id"])
        if prog:
            lines[0] += f" (برنامه: {prog['name']})"
    lines += [
        f"📅 {ju.jalali_pretty(expense['jalali_date'])}   💰 {ju.format_money(expense['amount'])}",
    ]
    for p in sorted(expense["participants"], key=lambda x: -x["share_amount"]):
        payer_tag = " (پرداخت‌کننده)" if p["user_id"] == expense["payer_id"] else ""
        lines.append(
            f"   • {p['first_name']}: {ju.format_money(p['share_amount'])}{payer_tag}"
        )
    if expense.get("receipt_file_id") or expense.get("receipt_text"):
        lines.append("   🧾 رسید ضمیمه داره")
    return "\n".join(lines)


async def _send_history_page(update_or_query, context, offset: int, edit: bool = False):
    db = get_db(context)
    requester = db.get_user_by_telegram_id(
        update_or_query.from_user.id
        if hasattr(update_or_query, "from_user")
        else update_or_query.effective_user.id
    )
    total = db.count_expenses()
    expenses = db.list_expenses(limit=HISTORY_PAGE_SIZE, offset=offset)

    if not expenses:
        text = "هیچ هزینه‌ای هنوز ثبت نشده."
        markup = None
    else:
        blocks = [_format_expense_detail(db, e) for e in expenses]
        header = f"🧾 <b>تاریخچه هزینه‌ها</b> ({offset + 1}-{offset + len(expenses)} از {total})\n"
        text = header + "\n\n".join(blocks)
        deletable_ids = [
            e["id"]
            for e in expenses
            if requester
            and (requester["is_admin"] or requester["id"] == e["creator_id"])
        ]
        receipt_ids = [
            e["id"]
            for e in expenses
            if e.get("receipt_file_id") or e.get("receipt_text")
        ]
        has_more = (offset + HISTORY_PAGE_SIZE) < total
        markup = kb.history_page_keyboard(offset, has_more, deletable_ids, receipt_ids)

    if edit:
        await update_or_query.edit_message_text(
            text, parse_mode="HTML", reply_markup=markup
        )
    else:
        await update_or_query.effective_message.reply_text(
            text, parse_mode="HTML", reply_markup=markup
        )


async def history_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = await get_active_user_or_warn(update, context)
    if user is None:
        return
    await _send_history_page(update, context, offset=0, edit=False)


async def history_page_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    offset = int(query.data.split("_", 1)[1])
    await _send_history_page(query, context, offset=offset, edit=True)


async def delete_expense_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    db = get_db(context)
    requester = db.get_user_by_telegram_id(update.effective_user.id)
    expense_id = int(query.data.split("_", 1)[1])
    expense = db.get_expense(expense_id)

    if expense is None:
        await query.answer("پیدا نشد.", show_alert=True)
        return
    if not requester or not (
        requester["is_admin"] or requester["id"] == expense["creator_id"]
    ):
        await query.answer(
            "فقط مدیر گروه یا کسی که این هزینه رو ثبت کرده می‌تونه حذفش کنه.",
            show_alert=True,
        )
        return

    db.delete_expense(expense_id)
    await query.answer("حذف شد ✅")
    await _send_history_page(query, context, offset=0, edit=True)


async def view_receipt_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    db = get_db(context)
    expense_id = int(query.data.split("_", 1)[1])
    expense = db.get_expense(expense_id)
    if expense is None:
        await query.message.reply_text("این هزینه دیگه پیدا نشد.")
        return
    if expense.get("receipt_file_id"):
        await context.bot.send_photo(
            chat_id=query.message.chat_id,
            photo=expense["receipt_file_id"],
            caption=f"🧾 رسید هزینه #{expense_id} — {expense['description']}",
        )
    elif expense.get("receipt_text"):
        await query.message.reply_text(
            f"🧾 توضیح رسید هزینه #{expense_id}:\n\n{expense['receipt_text']}"
        )
    else:
        await query.message.reply_text("این هزینه رسیدی ضمیمه نداره.")


# ------------------------------------------------------------- group report
async def group_report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin = await get_admin_or_warn(update, context)
    if admin is None:
        return
    db = get_db(context)
    pairs, simplified = group_report(db)

    lines = ["📋 <b>گزارش کامل بدهی‌های گروه</b>", "", "وضعیت فعلی (تهاتر شده):"]
    if not pairs:
        lines.append("  همه چی صافه ✅")
    else:
        for debtor, creditor, amt in pairs:
            lines.append(
                f"  • {debtor['first_name']} به {creditor['first_name']}: {ju.format_money(amt)}"
            )

    lines.append("")
    lines.append("💡 پیشنهاد تسویه گروهی با کمترین تعداد واریزی:")
    if not simplified:
        lines.append("  نیازی به هیچ واریزی نیست ✅")
    else:
        for frm, to, amt in simplified:
            card = (
                ju.format_card_number_html(to["card_number"])
                if to["card_number"]
                else "بدون شماره کارت"
            )
            lines.append(
                f"  • {frm['first_name']} ⬅️ {to['first_name']}: {ju.format_money(amt)}  (کارت: {card})"
            )

    lines.append("")
    lines.append("برای فایل اکسل کامل با ریز هزینه‌ها و پرداخت‌ها: /export")

    await update.effective_message.reply_text("\n".join(lines), parse_mode="HTML")


# --------------------------------------------------------------- xlsx export
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin = await get_admin_or_warn(update, context)
    if admin is None:
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db(context)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "هزینه‌ها"
    ws1.sheet_view.rightToLeft = True
    ws1.append(
        [
            "شماره",
            "تاریخ",
            "بابت",
            "مبلغ کل (تومان)",
            "پرداخت‌کننده",
            "نحوه تقسیم",
            "برنامه",
            "سهم هرکس",
            "رسید",
        ]
    )
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for e in db.list_expenses(limit=1_000_000):
        payer = db.get_user_by_id(e["payer_id"])
        shares_text = "، ".join(
            f"{p['first_name']}: {p['share_amount']:,}" for p in e["participants"]
        )
        mode_label = (
            "بر اساس تعداد نفرات"
            if e["split_mode"] == "weighted"
            else ("مساوی" if e["split_mode"] == "equal" else "مبلغ دلخواه")
        )
        prog_name = ""
        if e.get("program_id"):
            prog = db.get_program(e["program_id"])
            if prog:
                prog_name = prog["name"]
        has_receipt = (
            "دارد" if (e.get("receipt_file_id") or e.get("receipt_text")) else "ندارد"
        )
        ws1.append(
            [
                e["id"],
                e["jalali_date"],
                e["description"],
                e["amount"],
                payer["first_name"],
                mode_label,
                prog_name,
                shares_text,
                has_receipt,
            ]
        )

    ws2 = wb.create_sheet("پرداخت‌ها")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["شماره", "تاریخ", "از", "به", "مبلغ (تومان)", "وضعیت", "توضیح"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    status_fa = {
        "pending": "در انتظار تایید",
        "confirmed": "تایید شده",
        "rejected": "رد شده",
    }
    for p in db.list_payments(limit=1_000_000):
        frm = db.get_user_by_id(p["from_user_id"])
        to = db.get_user_by_id(p["to_user_id"])
        ws2.append(
            [
                p["id"],
                p["jalali_date"],
                frm["first_name"],
                to["first_name"],
                p["amount"],
                status_fa.get(p["status"], p["status"]),
                p["note"] or "",
            ]
        )

    ws3 = wb.create_sheet("وضعیت فعلی")
    ws3.sheet_view.rightToLeft = True
    ws3.append(["بدهکار", "طلبکار", "مبلغ (تومان)"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    pairs, simplified = group_report(db)
    for debtor, creditor, amt in pairs:
        ws3.append([debtor["first_name"], creditor["first_name"], amt])
    ws3.append([])
    ws3.append(["پیشنهاد تسویه با کمترین واریزی:"])
    for frm, to, amt in simplified:
        ws3.append([frm["first_name"], to["first_name"], amt])

    for ws in (ws1, ws2, ws3):
        for col_cells in ws.columns:
            length = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                45, max(12, length + 2)
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"hesabkitab_{ju.jalali_date_str().replace('/', '-')}.xlsx"
    await update.effective_message.reply_document(
        document=buffer, filename=filename, caption="📊 گزارش کامل حساب‌کتاب گروه"
    )


balance_command_handler = CommandHandler("balance", balance_command)
balance_button_handler = MessageHandler(
    filters.Regex(f"^{re.escape(kb.BTN_MY_BALANCE)}$"), balance_command
)

history_command_handler = CommandHandler("history", history_command)
history_button_handler = MessageHandler(
    filters.Regex(f"^{re.escape(kb.BTN_HISTORY)}$"), history_command
)
history_page_query_handler = CallbackQueryHandler(
    history_page_callback, pattern=r"^histpage_\d+$"
)
delete_expense_query_handler = CallbackQueryHandler(
    delete_expense_callback, pattern=r"^delexp_\d+$"
)
view_receipt_query_handler = CallbackQueryHandler(
    view_receipt_callback, pattern=r"^viewreceipt_\d+$"
)

report_command_handler = CommandHandler("report", group_report_command)
report_button_handler = MessageHandler(
    filters.Regex(f"^{re.escape(kb.BTN_ADMIN_REPORT)}$"), group_report_command
)

export_command_handler = CommandHandler("export", export_command)
